import openpyxl

wb = openpyxl.load_workbook('Etat_conversation_Marate.xlsx')
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]
print("Headers:")
for i, h in enumerate(headers):
    if h:
        print(f"  {i}: {h}")

# Get all rows
print("\nData (all rows):")
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
    if any(cell for cell in row):  # Only print non-empty rows
        print(f"Row {row_idx}: {row[:9]}")
